"""
Frontend router — serves Jinja2 HTML pages.
Auth is stored in an HttpOnly cookie (JWT token).
"""
from __future__ import annotations

import io
import json
import logging
import os
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from typing import List, Optional

from fastapi import (
    APIRouter, Depends, File, Form, HTTPException,
    Query, Request, Response, UploadFile,
)
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from jose import JWTError, jwt
from sqlalchemy import extract, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from src.core.auth import create_access_token, hash_password, verify_password
from src.core.database import get_db
from src.helpers import s3
from src.models.audit_log import EventAuditLog
from src.models.bookmark import Bookmark
from src.models.city import City
from src.models.country import Country
from src.models.event import Event
from src.models.organizer import Organizer
from src.models.tag import EventTag, Tag
from src.models.user import User

logger = logging.getLogger(__name__)
router = APIRouter(include_in_schema=False)

# ── Templates ──────────────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.dirname(__file__))  # src/
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

# ── JWT settings ───────────────────────────────────────────────────
SECRET_KEY = os.getenv("SECRET_KEY", "change-me-in-production")
ALGORITHM  = "HS256"
COOKIE_KEY = "access_token"


# ── Validation helpers ─────────────────────────────────────────────
def _strip(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    v = value.strip()
    return v if v else None


def _require(value: Optional[str], label: str) -> str:
    v = _strip(value)
    if not v:
        raise ValueError(f"«{label}» is required and cannot be blank.")
    return v


def _require_url(value: Optional[str], label: str) -> str:
    v = _require(value, label)
    if not (v.startswith("http://") or v.startswith("https://")):
        raise ValueError(f"«{label}» must be a valid URL starting with https://")
    return v


# ── Auth helpers ───────────────────────────────────────────────────
def _get_user_from_cookie(request: Request, db: AsyncSession):
    token = request.cookies.get(COOKIE_KEY)
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return int(payload["sub"]), payload.get("role", "user")
    except (JWTError, KeyError, ValueError):
        return None


async def get_current_user_cookie(
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> Optional[User]:
    info = _get_user_from_cookie(request, db)
    if not info:
        return None
    user_id, _ = info
    return await db.get(User, user_id)


def _set_auth_cookie(response: Response, token: str) -> None:
    response.set_cookie(
        COOKIE_KEY, token,
        httponly=True, samesite="lax",
        max_age=60 * 60 * 24,
    )


def _clear_auth_cookie(response: Response) -> None:
    response.delete_cookie(COOKIE_KEY)


def _require_admin(user: Optional[User]) -> None:
    if not user or user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")


def _resolve_tags_and_images(events):
    if not isinstance(events, list):
        events = [events]
    for ev in events:
        ev.tags = [et.tag for et in ev.event_tags]
        ev.image_url = s3.presign(ev.image_url)
    return events


# ── _ctx: does NOT include 'request' (passed separately per new Starlette API)
def _ctx(user: Optional[User], **kwargs) -> dict:
    return {"user": user, **kwargs}


def _r(request: Request, template: str, user: Optional[User],
        status_code: int = 200, **kwargs):
    """TemplateResponse wrapper compatible with Starlette 0.36+."""
    return templates.TemplateResponse(
        request, template, _ctx(user, **kwargs), status_code=status_code
    )


# ═══════════════════════════════════════════════════════════════════
# PUBLIC PAGES
# ═══════════════════════════════════════════════════════════════════

@router.get("/", response_class=HTMLResponse)
async def page_index(
    request: Request,
    search: Optional[str] = Query(None),
    is_online: Optional[str] = Query(None),
    organizer_id: Optional[str] = Query(None),
    tag_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _organizer_id = int(organizer_id) if organizer_id and organizer_id.strip() else None
    _tag_id       = int(tag_id)       if tag_id       and tag_id.strip()       else None
    query = select(Event).options(
        selectinload(Event.event_tags).selectinload(EventTag.tag)
    )
    if search:
        query = query.where(
            Event.title.ilike(f"%{search}%") | Event.description.ilike(f"%{search}%")
        )
    if is_online == "true":
        query = query.where(Event.is_online == True)
    elif is_online == "false":
        query = query.where(Event.is_online == False)
    if _organizer_id:
        query = query.where(Event.organizer_id == _organizer_id)
    if _tag_id:
        query = query.join(EventTag, Event.id == EventTag.event_id).where(EventTag.tag_id == _tag_id)

    query  = query.order_by(Event.date_start.asc())
    events = (await db.execute(query)).scalars().all()
    _resolve_tags_and_images(list(events))

    organizers       = (await db.execute(select(Organizer).order_by(Organizer.name))).scalars().all()
    tags             = (await db.execute(select(Tag).order_by(Tag.name))).scalars().all()
    total_events     = (await db.execute(select(func.count()).select_from(Event))).scalar()
    total_organizers = (await db.execute(select(func.count()).select_from(Organizer))).scalar()

    return _r(request, "index.html", user,
              events=events, organizers=organizers, tags=tags,
              search=search, is_online=is_online,
              organizer_id=_organizer_id, tag_id=_tag_id,
              total_events=total_events, total_organizers=total_organizers)


@router.get("/event/{event_id}", response_class=HTMLResponse)
async def page_event_detail(
    event_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    result = await db.execute(
        select(Event)
        .where(Event.id == event_id)
        .options(selectinload(Event.event_tags).selectinload(EventTag.tag))
    )
    event = result.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    _resolve_tags_and_images(event)
    organizer = await db.get(Organizer, event.organizer_id)

    is_bookmarked = False
    if user:
        bm = (await db.execute(
            select(Bookmark).where(
                Bookmark.user_id == user.id, Bookmark.event_id == event_id
            )
        )).scalar_one_or_none()
        is_bookmarked = bm is not None

    return _r(request, "event_detail.html", user,
               event=event, organizer=organizer, is_bookmarked=is_bookmarked)


# ═══════════════════════════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════════════════════════

@router.get("/login", response_class=HTMLResponse)
async def page_login(request: Request, user: Optional[User] = Depends(get_current_user_cookie)):
    if user:
        return RedirectResponse("/", status_code=302)
    return _r(request, "login.html", user)


@router.post("/login")
async def do_login(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    email = email.strip()
    if not email:
        return _r(request, "login.html", None, status_code=400, error="Email is required.")

    user = (await db.execute(select(User).where(User.email == email))).scalar_one_or_none()
    if not user or not verify_password(password, user.password_hash):
        return _r(request, "login.html", None, status_code=400,
                  error="Invalid email or password.", email=email)

    token    = create_access_token({"sub": str(user.id), "role": user.role})
    response = RedirectResponse("/", status_code=302)
    _set_auth_cookie(response, token)
    return response


@router.get("/register", response_class=HTMLResponse)
async def page_register(request: Request, user: Optional[User] = Depends(get_current_user_cookie)):
    if user:
        return RedirectResponse("/", status_code=302)
    return _r(request, "register.html", user)


@router.post("/register")
async def do_register(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    email = email.strip()
    if not email:
        return _r(request, "register.html", None, status_code=400, error="Email is required.")
    if not password or len(password) < 8:
        return _r(request, "register.html", None, status_code=400,
                  error="Password must be at least 8 characters.", email=email)

    existing = (await db.execute(select(User).where(User.email == email))).scalar_one_or_none()
    if existing:
        return _r(request, "register.html", None, status_code=400,
                  error="Email already registered.", email=email)

    new_user = User(email=email, password_hash=hash_password(password), role="user")
    db.add(new_user)
    await db.commit()

    token    = create_access_token({"sub": str(new_user.id), "role": new_user.role})
    response = RedirectResponse("/", status_code=302)
    _set_auth_cookie(response, token)
    return response


@router.get("/logout")
async def do_logout():
    response = RedirectResponse("/", status_code=302)
    _clear_auth_cookie(response)
    return response


# ═══════════════════════════════════════════════════════════════════
# BOOKMARKS
# ═══════════════════════════════════════════════════════════════════

@router.get("/bookmarks", response_class=HTMLResponse)
async def page_bookmarks(
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    if not user:
        return RedirectResponse("/login", status_code=302)

    result = await db.execute(
        select(Bookmark)
        .where(Bookmark.user_id == user.id)
        .options(
            selectinload(Bookmark.event)
            .selectinload(Event.event_tags)
            .selectinload(EventTag.tag)
        )
        .order_by(Bookmark.added_at.desc())
    )
    bookmarks = result.scalars().all()
    for bm in bookmarks:
        _resolve_tags_and_images(bm.event)

    return _r(request, "bookmarks.html", user, bookmarks=bookmarks)


@router.post("/bookmarks/{event_id}")
async def add_bookmark(
    event_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    if not user:
        return RedirectResponse("/login", status_code=302)

    existing = (await db.execute(
        select(Bookmark).where(Bookmark.user_id == user.id, Bookmark.event_id == event_id)
    )).scalar_one_or_none()

    if not existing:
        db.add(Bookmark(user_id=user.id, event_id=event_id))
        await db.commit()

    return RedirectResponse(f"/event/{event_id}", status_code=302)


@router.post("/bookmarks/{event_id}/remove")
async def remove_bookmark(
    event_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    if not user:
        return RedirectResponse("/login", status_code=302)

    bm = (await db.execute(
        select(Bookmark).where(Bookmark.user_id == user.id, Bookmark.event_id == event_id)
    )).scalar_one_or_none()

    if bm:
        await db.delete(bm)
        await db.commit()

    ref = request.headers.get("referer", "/bookmarks")
    return RedirectResponse(ref, status_code=302)


# ═══════════════════════════════════════════════════════════════════
# ADMIN — Dashboard
# ═══════════════════════════════════════════════════════════════════

@router.get("/admin", response_class=HTMLResponse)
async def page_admin_dashboard(
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    stats = {
        "events":     (await db.execute(select(func.count()).select_from(Event))).scalar(),
        "organizers": (await db.execute(select(func.count()).select_from(Organizer))).scalar(),
        "tags":       (await db.execute(select(func.count()).select_from(Tag))).scalar(),
        "cities":     (await db.execute(select(func.count()).select_from(City))).scalar(),
        "countries":  (await db.execute(select(func.count()).select_from(Country))).scalar(),
        "users":      (await db.execute(select(func.count()).select_from(User))).scalar(),
    }
    return _r(request, "admin/dashboard.html", user, stats=stats, active="dashboard")


# ═══════════════════════════════════════════════════════════════════
# ADMIN — Events
# ═══════════════════════════════════════════════════════════════════

@router.get("/admin/events", response_class=HTMLResponse)
async def page_admin_events(
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    events = (await db.execute(
        select(Event)
        .options(selectinload(Event.event_tags).selectinload(EventTag.tag))
        .order_by(Event.id.desc())
    )).scalars().all()
    for ev in events:
        ev.tags = [et.tag for et in ev.event_tags]

    return _r(request, "admin/events.html", user, events=events, active="events")


@router.get("/admin/events/new", response_class=HTMLResponse)
async def page_admin_event_new(
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    organizers = (await db.execute(select(Organizer).order_by(Organizer.name))).scalars().all()
    cities     = (await db.execute(select(City).order_by(City.name))).scalars().all()
    tags       = (await db.execute(select(Tag).order_by(Tag.name))).scalars().all()
    return _r(request, "admin/event_form.html", user,
               event=None, organizers=organizers, cities=cities, tags=tags)


async def _event_form_error(request, user, db, event, error, status_code=400):
    organizers = (await db.execute(select(Organizer).order_by(Organizer.name))).scalars().all()
    cities     = (await db.execute(select(City).order_by(City.name))).scalars().all()
    tags       = (await db.execute(select(Tag).order_by(Tag.name))).scalars().all()
    return _r(request, "admin/event_form.html", user, status_code=status_code,
               event=event, error=error, organizers=organizers, cities=cities, tags=tags)


@router.post("/admin/events/new")
async def do_admin_event_create(
    request: Request,
    title: str = Form(...),
    organizer_id: int = Form(...),
    date_start: datetime = Form(...),
    date_end: Optional[datetime] = Form(None),
    registration_deadline: Optional[datetime] = Form(None),
    price: Decimal = Form(Decimal("0.00")),
    city_id: Optional[int] = Form(None),
    location_address: Optional[str] = Form(None),
    website_url: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    is_online: Optional[str] = Form(None),
    tag_ids: List[int] = Form([]),
    image: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    try:
        title       = _require(title, "Title")
        description = _require(description, "Description")
        website_url = _require_url(website_url, "Website URL")
    except ValueError as exc:
        return await _event_form_error(request, user, db, None, str(exc))

    online = (is_online == "true")
    if not online:
        location_address = _strip(location_address)
        if not location_address:
            return await _event_form_error(
                request, user, db, None,
                "«Location address» is required for in-person events."
            )

    event = Event(
        title=title, organizer_id=organizer_id,
        date_start=date_start, date_end=date_end,
        registration_deadline=registration_deadline,
        price=price, city_id=city_id or None,
        location_address=location_address,
        website_url=website_url, description=description,
        is_online=online,
    )

    if image and image.filename:
        if image.content_type not in s3.ALLOWED_IMG:
            return await _event_form_error(request, user, db, None, "Unsupported image type.")
        contents = await image.read()
        if contents:
            ext    = s3.ALLOWED_IMG[image.content_type]
            s3_key = f"uploads/{uuid.uuid4()}.{ext}"
            try:
                s3.upload_fileobj(io.BytesIO(contents), s3_key, image.content_type)
                event.image_url = s3_key
            except Exception as e:
                logger.error("S3 upload failed: %s", e)

    db.add(event)
    await db.flush()
    for tid in tag_ids:
        db.add(EventTag(event_id=event.id, tag_id=tid))
    await db.commit()
    return RedirectResponse("/admin/events", status_code=302)


@router.get("/admin/events/{event_id}/edit", response_class=HTMLResponse)
async def page_admin_event_edit(
    event_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    result = await db.execute(
        select(Event)
        .where(Event.id == event_id)
        .options(selectinload(Event.event_tags).selectinload(EventTag.tag))
    )
    event = result.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404)

    event.tags      = [et.tag for et in event.event_tags]
    event.image_url = s3.presign(event.image_url)

    organizers = (await db.execute(select(Organizer).order_by(Organizer.name))).scalars().all()
    cities     = (await db.execute(select(City).order_by(City.name))).scalars().all()
    tags       = (await db.execute(select(Tag).order_by(Tag.name))).scalars().all()
    return _r(request, "admin/event_form.html", user,
               event=event, organizers=organizers, cities=cities, tags=tags)


@router.post("/admin/events/{event_id}/edit")
async def do_admin_event_edit(
    event_id: int,
    request: Request,
    title: str = Form(...),
    organizer_id: int = Form(...),
    date_start: datetime = Form(...),
    date_end: Optional[datetime] = Form(None),
    registration_deadline: Optional[datetime] = Form(None),
    price: Decimal = Form(Decimal("0.00")),
    city_id: Optional[int] = Form(None),
    location_address: Optional[str] = Form(None),
    website_url: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    is_online: Optional[str] = Form(None),
    tag_ids: List[int] = Form([]),
    image: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)

    async def _fetch_event_for_form():
        res = await db.execute(
            select(Event).where(Event.id == event_id)
            .options(selectinload(Event.event_tags).selectinload(EventTag.tag))
        )
        ev = res.scalar_one_or_none()
        if ev:
            ev.tags      = [et.tag for et in ev.event_tags]
            ev.image_url = s3.presign(ev.image_url)
        return ev

    try:
        title       = _require(title, "Title")
        description = _require(description, "Description")
        website_url = _require_url(website_url, "Website URL")
    except ValueError as exc:
        ev = await _fetch_event_for_form()
        return await _event_form_error(request, user, db, ev, str(exc))

    online = (is_online == "true")
    if not online:
        location_address = _strip(location_address)
        if not location_address:
            ev = await _fetch_event_for_form()
            return await _event_form_error(
                request, user, db, ev,
                "«Location address» is required for in-person events."
            )

    result = await db.execute(
        select(Event)
        .where(Event.id == event_id)
        .options(selectinload(Event.event_tags).selectinload(EventTag.tag))
    )
    event = result.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404)

    updates = {
        "title": title, "organizer_id": organizer_id,
        "date_start": date_start, "date_end": date_end,
        "registration_deadline": registration_deadline,
        "price": price, "city_id": city_id or None,
        "location_address": location_address,
        "website_url": website_url, "description": description,
        "is_online": online,
    }

    def _fmt(v):
        if isinstance(v, datetime):
            return v.replace(tzinfo=None).strftime("%Y-%m-%d %H:%M:%S")
        if v is None:
            return ""
        return str(v)

    for field, new_val in updates.items():
        old_val = getattr(event, field, None)
        old_str, new_str = _fmt(old_val), _fmt(new_val)
        if old_str != new_str:
            db.add(EventAuditLog(
                event_id=event_id, changed_by=user.id,
                changed_column=field, old_value=old_str, new_value=new_str,
            ))
        setattr(event, field, new_val)

    if image and image.filename and image.content_type in s3.ALLOWED_IMG:
        contents = await image.read()
        if contents:
            ext    = s3.ALLOWED_IMG[image.content_type]
            s3_key = f"uploads/{uuid.uuid4()}.{ext}"
            try:
                s3.upload_fileobj(io.BytesIO(contents), s3_key, image.content_type)
                event.image_url = s3_key
            except Exception as e:
                logger.error("S3 upload failed: %s", e)

    await db.execute(EventTag.__table__.delete().where(EventTag.event_id == event_id))
    for tid in tag_ids:
        db.add(EventTag(event_id=event_id, tag_id=tid))

    event.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return RedirectResponse("/admin/events", status_code=302)


@router.post("/admin/events/{event_id}/delete")
async def do_admin_event_delete(
    event_id: int,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    event = await db.get(Event, event_id)
    if event:
        s3.delete_object(event.image_url)
        await db.delete(event)
        await db.commit()
    return RedirectResponse("/admin/events", status_code=302)


# ═══════════════════════════════════════════════════════════════════
# ADMIN — Tags
# ═══════════════════════════════════════════════════════════════════

@router.get("/admin/tags", response_class=HTMLResponse)
async def page_admin_tags(
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    tags = (await db.execute(select(Tag).order_by(Tag.name))).scalars().all()
    return _r(request, "admin/tags.html", user, tags=tags, active="tags")


@router.post("/admin/tags")
async def do_admin_tag_create(
    request: Request,
    name: str = Form(...),
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    name = _strip(name)
    if not name:
        tags = (await db.execute(select(Tag).order_by(Tag.name))).scalars().all()
        return _r(request, "admin/tags.html", user, status_code=400,
                   tags=tags, error="Tag name cannot be blank.")

    existing = (await db.execute(select(Tag).where(Tag.name == name))).scalar_one_or_none()
    if existing:
        tags = (await db.execute(select(Tag).order_by(Tag.name))).scalars().all()
        return _r(request, "admin/tags.html", user, status_code=400,
                   tags=tags, error=f"Tag '{name}' already exists.")

    db.add(Tag(name=name))
    await db.commit()
    return RedirectResponse("/admin/tags", status_code=302)


@router.post("/admin/tags/{tag_id}/delete")
async def do_admin_tag_delete(
    tag_id: int,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    tag = await db.get(Tag, tag_id)
    if tag:
        await db.delete(tag)
        await db.commit()
    return RedirectResponse("/admin/tags", status_code=302)


# ═══════════════════════════════════════════════════════════════════
# ADMIN — Cities
# ═══════════════════════════════════════════════════════════════════

@router.get("/admin/cities", response_class=HTMLResponse)
async def page_admin_cities(
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    cities    = (await db.execute(
        select(City).options(selectinload(City.country)).order_by(City.name)
    )).scalars().all()
    countries = (await db.execute(select(Country).order_by(Country.name))).scalars().all()
    return _r(request, "admin/cities.html", user,
               cities=cities, countries=countries, active="cities")


@router.post("/admin/cities")
async def do_admin_city_create(
    request: Request,
    name: str = Form(...),
    country_id: int = Form(...),
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)

    async def _city_error(msg: str):
        cities    = (await db.execute(
            select(City).options(selectinload(City.country)).order_by(City.name)
        )).scalars().all()
        countries = (await db.execute(select(Country).order_by(Country.name))).scalars().all()
        return _r(request, "admin/cities.html", user, status_code=400,
                   cities=cities, countries=countries, active="cities", error=msg)

    name = _strip(name)
    if not name:
        return await _city_error("City name cannot be blank.")

    country = await db.get(Country, country_id)
    if not country:
        return await _city_error("Please select a valid country.")

    existing = (await db.execute(
        select(City).where(City.name == name, City.country_id == country_id)
    )).scalar_one_or_none()
    if existing:
        return await _city_error(f"City '{name}' already exists in {country.name}.")

    db.add(City(name=name, country_id=country_id))
    await db.commit()
    return RedirectResponse("/admin/cities", status_code=302)


@router.post("/admin/cities/{city_id}/delete")
async def do_admin_city_delete(
    city_id: int,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    city = await db.get(City, city_id)
    if city:
        await db.delete(city)
        await db.commit()
    return RedirectResponse("/admin/cities", status_code=302)


# ═══════════════════════════════════════════════════════════════════
# ADMIN — Organizers
# ═══════════════════════════════════════════════════════════════════

@router.get("/admin/organizers", response_class=HTMLResponse)
async def page_admin_organizers(
    request: Request,
    error: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    organizers = (await db.execute(select(Organizer).order_by(Organizer.name))).scalars().all()
    for org in organizers:
        org.event_count = (await db.execute(
            select(func.count()).select_from(Event).where(Event.organizer_id == org.id)
        )).scalar()
    return _r(request, "admin/organizers.html", user,
               organizers=organizers, active="organizers", error=error)


@router.post("/admin/organizers")
async def do_admin_organizer_create(
    request: Request,
    name: str = Form(...),
    website: Optional[str] = Form(None),
    contact_email: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)

    async def _org_error(msg: str):
        orgs = (await db.execute(select(Organizer).order_by(Organizer.name))).scalars().all()
        return _r(request, "admin/organizers.html", user, status_code=400,
                   organizers=orgs, error=msg)

    try:
        name          = _require(name, "Name")
        website       = _require_url(website, "Website")
        contact_email = _require(contact_email, "Contact email")
        description   = _require(description, "Description")
    except ValueError as exc:
        return await _org_error(str(exc))

    if "@" not in contact_email:
        return await _org_error("«Contact email» must be a valid email address.")

    existing = (await db.execute(select(Organizer).where(Organizer.name == name))).scalar_one_or_none()
    if existing:
        return await _org_error(f"Organizer '{name}' already exists.")

    db.add(Organizer(name=name, website=website, contact_email=contact_email, description=description))
    await db.commit()
    return RedirectResponse("/admin/organizers", status_code=302)


@router.post("/admin/organizers/{organizer_id}/delete")
async def do_admin_organizer_delete(
    organizer_id: int,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)

    organizer = await db.get(Organizer, organizer_id)
    if not organizer:
        return RedirectResponse("/admin/organizers", status_code=302)

    linked = (await db.execute(
        select(func.count()).select_from(Event).where(Event.organizer_id == organizer_id)
    )).scalar()

    if linked:
        from urllib.parse import quote
        msg = quote(f'Organizer "{organizer.name}" has {linked} event(s) and cannot be deleted.')
        return RedirectResponse(f"/admin/organizers?error={msg}", status_code=302)

    await db.delete(organizer)
    await db.commit()
    return RedirectResponse("/admin/organizers", status_code=302)


# ═══════════════════════════════════════════════════════════════════
# ADMIN — Audit Log
# ═══════════════════════════════════════════════════════════════════

@router.get("/admin/audit", response_class=HTMLResponse)
async def page_admin_audit(
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    logs = (await db.execute(
        select(EventAuditLog).order_by(EventAuditLog.change_date.desc()).limit(200)
    )).scalars().all()

    user_ids = {log.changed_by for log in logs if log.changed_by is not None}
    users_by_id: dict[int, str] = {}
    if user_ids:
        rows = (await db.execute(
            select(User.id, User.email).where(User.id.in_(user_ids))
        )).all()
        users_by_id = {row.id: row.email for row in rows}

    return _r(request, "admin/audit.html", user,
               logs=logs, users_by_id=users_by_id, active="audit")


# ═══════════════════════════════════════════════════════════════════
# ADMIN — Countries
# ═══════════════════════════════════════════════════════════════════

@router.get("/admin/countries", response_class=HTMLResponse)
async def page_admin_countries(
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    countries = (await db.execute(
        select(Country).options(selectinload(Country.cities)).order_by(Country.name)
    )).scalars().all()
    return _r(request, "admin/countries.html", user, countries=countries, active="countries")


@router.post("/admin/countries")
async def do_admin_country_create(
    request: Request,
    name: str = Form(...),
    iso_code: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)

    async def _country_error(msg: str):
        countries = (await db.execute(
            select(Country).options(selectinload(Country.cities)).order_by(Country.name)
        )).scalars().all()
        return _r(request, "admin/countries.html", user, status_code=400,
                   countries=countries, active="countries", error=msg)

    try:
        name = _require(name, "Country name")
    except ValueError as exc:
        return await _country_error(str(exc))

    iso = _strip(iso_code)
    if iso:
        iso = iso.upper()
        if not (2 <= len(iso) <= 3 and iso.isalpha()):
            return await _country_error("ISO code must be 2–3 letters (e.g. DE, UKR).")
        existing_iso = (await db.execute(
            select(Country).where(Country.iso_code == iso)
        )).scalar_one_or_none()
        if existing_iso:
            return await _country_error(f"ISO code '{iso}' is already used by {existing_iso.name}.")

    existing_name = (await db.execute(
        select(Country).where(Country.name == name)
    )).scalar_one_or_none()
    if existing_name:
        return await _country_error(f"Country '{name}' already exists.")

    db.add(Country(name=name, iso_code=iso or None))
    await db.commit()
    return RedirectResponse("/admin/countries", status_code=302)


@router.post("/admin/countries/{country_id}/delete")
async def do_admin_country_delete(
    country_id: int,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)
    country = await db.get(Country, country_id)
    if country:
        await db.delete(country)
        await db.commit()
    return RedirectResponse("/admin/countries", status_code=302)


# ═══════════════════════════════════════════════════════════════════
# ADMIN — Statistics  (1.5 — Data Visualization)
# ═══════════════════════════════════════════════════════════════════

@router.get("/admin/stats", response_class=HTMLResponse)
async def page_admin_stats(
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)

    tag_rows = (await db.execute(
        select(Tag.name, func.count(EventTag.event_id).label("cnt"))
        .join(EventTag, Tag.id == EventTag.tag_id)
        .group_by(Tag.name)
        .order_by(func.count(EventTag.event_id).desc())
        .limit(10)
    )).all()

    org_rows = (await db.execute(
        select(Organizer.name, func.count(Event.id).label("cnt"))
        .join(Event, Organizer.id == Event.organizer_id)
        .group_by(Organizer.name)
        .order_by(func.count(Event.id).desc())
        .limit(10)
    )).all()

    online_count   = (await db.execute(select(func.count()).select_from(Event).where(Event.is_online == True))).scalar()  or 0
    inperson_count = (await db.execute(select(func.count()).select_from(Event).where(Event.is_online == False))).scalar() or 0
    free_count     = (await db.execute(select(func.count()).select_from(Event).where(Event.price == 0))).scalar() or 0
    paid_count     = (await db.execute(select(func.count()).select_from(Event).where(Event.price >  0))).scalar() or 0

    month_rows = (await db.execute(
        select(
            extract("year",  Event.created_at).label("yr"),
            extract("month", Event.created_at).label("mo"),
            func.count(Event.id).label("cnt"),
        )
        .group_by("yr", "mo")
        .order_by("yr", "mo")
    )).all()

    return _r(request, "admin/stats.html", user,
               active="stats",
               total_events=online_count + inperson_count,
               online_count=online_count,
               inperson_count=inperson_count,
               free_count=free_count,
               paid_count=paid_count,
               tag_labels=json.dumps([r.name for r in tag_rows]),
               tag_data=json.dumps([r.cnt for r in tag_rows]),
               org_labels=json.dumps([r.name for r in org_rows]),
               org_data=json.dumps([r.cnt for r in org_rows]),
               format_labels=json.dumps(["Online", "In-person"]),
               format_data=json.dumps([online_count, inperson_count]),
               price_labels=json.dumps(["Free", "Paid"]),
               price_data=json.dumps([free_count, paid_count]),
               monthly_labels=json.dumps([f"{int(r.yr)}-{int(r.mo):02d}" for r in month_rows]),
               monthly_data=json.dumps([r.cnt for r in month_rows]))


# ═══════════════════════════════════════════════════════════════════
# ADMIN — Excel Export  (1.6 — File Operations, Export)
# ═══════════════════════════════════════════════════════════════════

@router.get("/admin/events/export")
async def export_events_xlsx(
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    events = (await db.execute(
        select(Event)
        .options(
            selectinload(Event.event_tags).selectinload(EventTag.tag),
            selectinload(Event.city),
            selectinload(Event.organizer),
        )
        .order_by(Event.id)
    )).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Events"

    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HDR_FILL  = PatternFill(fill_type="solid", start_color="1A1A2E")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALT_FILL  = PatternFill(fill_type="solid", start_color="F0F0F5")
    DATA_FONT = Font(name="Arial", size=10)

    columns = [
        ("ID", 6), ("Title", 30), ("Organizer", 22), ("Date Start", 18),
        ("Date End", 18), ("Price (USD)", 12), ("Format", 12),
        ("Location / City", 28), ("Website URL", 40), ("Description", 50),
        ("Tags", 28), ("Created At", 18),
    ]
    for col, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = HDR_ALIGN
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for row_idx, ev in enumerate(events, 2):
        tags_str = ", ".join(et.tag.name for et in ev.event_tags)
        location = ev.location_address or ""
        if not location and ev.city:
            location = ev.city.name
        if not location and ev.is_online:
            location = "Online"
        row_vals = [
            ev.id, ev.title,
            ev.organizer.name if ev.organizer else "",
            ev.date_start.strftime("%Y-%m-%d %H:%M") if ev.date_start else "",
            ev.date_end.strftime("%Y-%m-%d %H:%M")   if ev.date_end   else "",
            float(ev.price) if ev.price else 0.0,
            "Online" if ev.is_online else "In-person",
            location, ev.website_url or "", ev.description or "",
            tags_str,
            ev.created_at.strftime("%Y-%m-%d %H:%M") if ev.created_at else "",
        ]
        use_alt = (row_idx % 2 == 0)
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 10))
            if use_alt:
                cell.fill = ALT_FILL

    ws2 = wb.create_sheet("Import Template")
    TMPL_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    TMPL_FILL = PatternFill(fill_type="solid", start_color="C9A84C")
    import_cols = [
        ("title", 30), ("organizer_name", 22), ("date_start", 22), ("date_end", 22),
        ("price", 12), ("is_online", 14), ("location_address", 30),
        ("website_url", 40), ("description", 50), ("tags", 30), ("city_name", 20),
    ]
    for col, (header, width) in enumerate(import_cols, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = TMPL_FONT; cell.fill = TMPL_FILL
        cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions[get_column_letter(col)].width = width
    notes = [
        "Required. Max 100 chars", "Required. Must match existing organizer name exactly",
        "Required. Format: YYYY-MM-DD HH:MM", "Optional. Format: YYYY-MM-DD HH:MM",
        "Optional. Default: 0.00", "Optional. TRUE or FALSE (default FALSE)",
        "Required for in-person events", "Optional. Must start with https://",
        "Optional.", "Optional. Comma-separated tag names", "Optional. Must match existing city name",
    ]
    notes_font = Font(name="Arial", italic=True, color="6B6876", size=9)
    for col, note in enumerate(notes, 1):
        cell = ws2.cell(row=2, column=col, value=note); cell.font = notes_font
    sample = [
        "Sample Conference 2026", "IT Arena", "2026-09-10 09:00", "2026-09-12 18:00",
        "199.00", "FALSE", "123 Main St, Lviv", "https://example.com",
        "A great tech conference", "Backend, DevOps, Cloud", "Lviv",
    ]
    sf = Font(name="Arial", size=10, color="1A1A2E")
    sfill = PatternFill(fill_type="solid", start_color="FFFBE6")
    for col, val in enumerate(sample, 1):
        cell = ws2.cell(row=3, column=col, value=val); cell.font = sf; cell.fill = sfill
    ws2.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="events_export.xlsx"'},
    )


# ═══════════════════════════════════════════════════════════════════
# ADMIN — Excel Import  (1.6 — File Operations, Import)
# ═══════════════════════════════════════════════════════════════════

@router.post("/admin/events/import")
async def import_events_xlsx(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: Optional[User] = Depends(get_current_user_cookie),
):
    _require_admin(user)

    from urllib.parse import quote
    from openpyxl import load_workbook

    if not (file.filename or "").lower().endswith(".xlsx"):
        msg = quote("Import failed: only .xlsx files are supported.")
        return RedirectResponse(f"/admin/events?import_status=error&msg={msg}", status_code=302)

    contents = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    except Exception:
        msg = quote("Import failed: could not open the file.")
        return RedirectResponse(f"/admin/events?import_status=error&msg={msg}", status_code=302)

    ws = wb["Import Template"] if "Import Template" in wb.sheetnames else wb.worksheets[0]

    def clean_hdr(raw) -> str:
        h = str(raw or "").strip().lower()
        return h.split("*")[0].split("(")[0].strip()

    headers = [clean_hdr(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]

    def cell_str(row, col_name: str) -> str:
        if col_name not in headers:
            return ""
        return str(ws.cell(row=row, column=headers.index(col_name) + 1).value or "").strip()

    created_count = 0
    error_rows: list[str] = []

    for row_num in range(2, ws.max_row + 1):
        title = cell_str(row_num, "title")
        if not title:
            continue

        org_name = cell_str(row_num, "organizer_name")
        if not org_name:
            error_rows.append(f"Row {row_num} ({title!r}): organizer_name is required"); continue

        org = (await db.execute(select(Organizer).where(Organizer.name == org_name))).scalar_one_or_none()
        if not org:
            error_rows.append(f"Row {row_num} ({title!r}): organizer '{org_name}' not found"); continue

        ds_raw = ws.cell(row=row_num, column=headers.index("date_start") + 1).value \
            if "date_start" in headers else None
        try:
            date_start = ds_raw if isinstance(ds_raw, datetime) \
                else datetime.strptime(str(ds_raw).strip(), "%Y-%m-%d %H:%M")
        except (ValueError, TypeError):
            error_rows.append(f"Row {row_num} ({title!r}): invalid date_start '{ds_raw}'"); continue

        de_raw = ws.cell(row=row_num, column=headers.index("date_end") + 1).value \
            if "date_end" in headers else None
        date_end = None
        if de_raw:
            try:
                date_end = de_raw if isinstance(de_raw, datetime) \
                    else datetime.strptime(str(de_raw).strip(), "%Y-%m-%d %H:%M")
            except (ValueError, TypeError):
                pass

        try:
            price = Decimal(str(cell_str(row_num, "price") or "0"))
        except Exception:
            price = Decimal("0.00")

        is_online = cell_str(row_num, "is_online").upper() in ("TRUE", "1", "YES")

        city_id = None
        city_name = cell_str(row_num, "city_name")
        if city_name:
            city_obj = (await db.execute(
                select(City).where(City.name == city_name)
            )).scalar_one_or_none()
            if city_obj:
                city_id = city_obj.id

        ev = Event(
            title=title, organizer_id=org.id,
            date_start=date_start, date_end=date_end,
            price=price, is_online=is_online,
            location_address=cell_str(row_num, "location_address") or None,
            website_url=cell_str(row_num, "website_url") or None,
            description=cell_str(row_num, "description") or None,
            city_id=city_id,
        )
        db.add(ev)
        await db.flush()

        tags_str = cell_str(row_num, "tags")
        if tags_str:
            for tag_name in [t.strip() for t in tags_str.split(",") if t.strip()]:
                tag_obj = (await db.execute(
                    select(Tag).where(Tag.name == tag_name)
                )).scalar_one_or_none()
                if tag_obj:
                    db.add(EventTag(event_id=ev.id, tag_id=tag_obj.id))

        created_count += 1

    await db.commit()

    parts = [f"Successfully imported {created_count} event(s)."]
    if error_rows:
        parts.append("Skipped: " + " | ".join(error_rows[:5]))
        if len(error_rows) > 5:
            parts.append(f"... and {len(error_rows) - 5} more.")

    status = "ok" if created_count > 0 else "error"
    return RedirectResponse(
        f"/admin/events?import_status={status}&msg={quote(' '.join(parts))}",
        status_code=302,
    )